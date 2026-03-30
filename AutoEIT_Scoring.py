"""
AutoEIT GSoC 2026 - Test II: Automated Scoring System
Author: Anand S (sajithanand99@gmail.com)
Project: AutoEIT2 - Automated scoring system for EIT responses
Organization: HumanAI Foundation

Approach:
---------
This script implements a meaning-based rubric scoring engine for the
Spanish Elicited Imitation Task (EIT). The system:
1. Preprocesses both target (stimulus) and learner transcription
2. Extracts content words (meaningful lexical items)
3. Computes fuzzy token-level overlap to handle spelling variations
4. Applies a 0-4 rubric based on proportion of meaning units reproduced
5. Writes sentence-level scores back to the Excel file

Rubric (Meaning-Based, 0-4 Scale):
------------------------------------
4 - Full meaning conveyed: >= 90% content word overlap, sentence largely intact
3 - Most meaning conveyed: 65-89% overlap, minor omissions/substitutions
2 - Partial meaning conveyed: 35-64% overlap, significant errors but core meaning present
1 - Minimal meaning conveyed: 10-34% overlap, heavily fragmented
0 - No meaning conveyed: <10% overlap OR fully unintelligible/empty response
"""

import re
import unicodedata
import openpyxl
import pandas as pd
from rapidfuzz import fuzz
from copy import copy

# ─────────────────────────────────────────────
# SECTION 1: SPANISH FUNCTION WORD LIST
# These are excluded from content word extraction
# ─────────────────────────────────────────────

SPANISH_STOPWORDS = {
    'a', 'al', 'ante', 'bajo', 'con', 'contra', 'de', 'del', 'desde',
    'durante', 'en', 'entre', 'hacia', 'hasta', 'mediante', 'para',
    'por', 'según', 'sin', 'sobre', 'tras',
    'el', 'la', 'los', 'las', 'un', 'una', 'unos', 'unas',
    'y', 'e', 'ni', 'o', 'u', 'pero', 'sino', 'que', 'aunque',
    'si', 'porque', 'cuando', 'donde', 'como', 'mientras',
    'yo', 'tú', 'él', 'ella', 'nosotros', 'vosotros', 'ellos', 'ellas',
    'me', 'te', 'se', 'nos', 'os', 'le', 'les', 'lo', 'la', 'mi', 'mis',
    'su', 'sus', 'tu', 'tus', 'este', 'esta', 'esto', 'ese', 'esa', 'eso',
    'aquel', 'aquella', 'aquello',
    'es', 'son', 'está', 'están', 'ser', 'estar', 'fue', 'era',
    'hay', 'ha', 'han', 'he', 'haber',
    'no', 'sí', 'ya', 'más', 'muy', 'tan', 'también', 'tampoco',
    'que', 'cual', 'quien', 'cuyo', 'cuanto',
}

# ─────────────────────────────────────────────
# SECTION 2: TEXT PREPROCESSING
# ─────────────────────────────────────────────

DISFLUENCY_PATTERNS = [
    r'\[.*?\]',           # [pause], [gibberish], [cough]
    r'\bxxx\b',           # xxx filler
    r'\bx\b',             # single x marker
    r'\(.*?\)',           # (parenthetical notes) -- but careful: stimulus has (word_count)
    r'\.{2,}',            # ellipsis ...
    r'\b\w+-\b',          # partial words like "fu-" "Mis gus.."
    r'\bum\b|\buh\b|\bmm\b|\bmhh?\b|\beh\b',  # English/Spanish fillers
]

def normalize_text(text):
    """Lowercase, remove accents for fuzzy matching, strip punctuation."""
    if not text:
        return ""
    text = text.lower()
    # Normalize unicode (remove accents for comparison)
    nfkd = unicodedata.normalize('NFKD', text)
    text_no_accent = ''.join(c for c in nfkd if not unicodedata.combining(c))
    # Remove punctuation except letters/numbers/spaces
    text_clean = re.sub(r"[^\w\s]", " ", text_no_accent)
    return re.sub(r'\s+', ' ', text_clean).strip()


def clean_stimulus(stimulus):
    """Remove word count annotation like (7), (12) from end of stimulus."""
    return re.sub(r'\s*\(\d+\)\s*$', '', stimulus).strip()


def clean_transcription(transcription):
    """Remove disfluencies, fillers, partial words, and annotations."""
    if not transcription:
        return ""
    text = str(transcription)
    for pattern in DISFLUENCY_PATTERNS:
        text = re.sub(pattern, ' ', text, flags=re.IGNORECASE)
    # Remove leftover punctuation (keep letters, spaces)
    text = re.sub(r"[^\w\s]", " ", text)
    text = re.sub(r'\s+', ' ', text).strip()
    return text


def extract_content_words(text):
    """Return list of content words (non-stopwords) from normalized text."""
    words = normalize_text(text).split()
    return [w for w in words if w not in SPANISH_STOPWORDS and len(w) > 1]


# ─────────────────────────────────────────────
# SECTION 3: FUZZY WORD MATCHING
# ─────────────────────────────────────────────

def fuzzy_word_overlap(target_words, response_words, threshold=80):
    """
    For each target content word, check if a sufficiently similar
    word exists in the response. Uses RapidFuzz token ratio.
    Returns: (matched_count, total_target_words, match_ratio)
    """
    if not target_words:
        return 0, 0, 0.0

    matched = 0
    used = set()  # prevent double-counting response words

    for tw in target_words:
        best_score = 0
        best_idx = -1
        for i, rw in enumerate(response_words):
            if i in used:
                continue
            score = fuzz.ratio(tw, rw)
            if score > best_score:
                best_score = score
                best_idx = i
        if best_score >= threshold:
            matched += 1
            used.add(best_idx)

    ratio = matched / len(target_words)
    return matched, len(target_words), ratio


# ─────────────────────────────────────────────
# SECTION 4: MEANING-BASED RUBRIC SCORING
# ─────────────────────────────────────────────

def is_unintelligible(transcription_raw):
    """
    Detect if response is essentially empty or fully unintelligible
    after cleaning disfluencies.
    """
    cleaned = clean_transcription(str(transcription_raw))
    # After cleaning, if barely any content left
    tokens = [t for t in cleaned.split() if len(t) > 1]
    return len(tokens) == 0


def score_sentence(stimulus, transcription):
    """
    Apply the meaning-based rubric and return a score 0-4.

    Rubric:
      4 → >= 90% content word overlap
      3 → 65–89% content word overlap
      2 → 35–64% content word overlap
      1 → 10–34% content word overlap
      0 → < 10% OR unintelligible/empty
    """
    # Handle missing/empty transcription
    if pd.isna(transcription) or str(transcription).strip() == '':
        return 0

    raw_transcription = str(transcription)

    # Early exit for unintelligible responses
    if is_unintelligible(raw_transcription):
        return 0

    # Clean both sides
    clean_stim = clean_stimulus(str(stimulus))
    clean_trans = clean_transcription(raw_transcription)

    # Extract content words
    target_words = extract_content_words(clean_stim)
    response_words = extract_content_words(clean_trans)

    if not target_words:
        return 0

    # Fuzzy overlap
    matched, total, ratio = fuzzy_word_overlap(target_words, response_words)

    # Apply rubric thresholds
    if ratio >= 0.90:
        return 4
    elif ratio >= 0.65:
        return 3
    elif ratio >= 0.35:
        return 2
    elif ratio >= 0.10:
        return 1
    else:
        return 0


# ─────────────────────────────────────────────
# SECTION 5: PROCESS ALL PARTICIPANTS
# ─────────────────────────────────────────────

def process_workbook(input_path, output_path):
    """
    Read Excel, apply scoring to each participant sheet,
    write scores to the 'Score' column, save output.
    """
    wb = openpyxl.load_workbook(input_path)
    participant_sheets = [s for s in wb.sheetnames if s != 'Info']

    results_summary = {}

    for sheet_name in participant_sheets:
        ws = wb[sheet_name]
        print(f"\n{'='*55}")
        print(f" Participant: {sheet_name}")
        print(f"{'='*55}")
        print(f"{'#':<5} {'Score':<7} {'Match%':<9} {'Stimulus (cleaned)':<35} {'Transcription (cleaned)'}")
        print("-" * 120)

        scores = []
        # Find Score column (column D = index 4)
        score_col = 4

        for row in ws.iter_rows(min_row=2):
            sentence_no = row[0].value
            stimulus = row[1].value
            transcription = row[2].value

            if sentence_no is None:
                continue

            # Score it
            score = score_sentence(stimulus, transcription)
            row[score_col - 1].value = score
            scores.append(score)

            # For display
            clean_s = clean_stimulus(str(stimulus))[:33] if stimulus else ''
            clean_t = clean_transcription(str(transcription))[:45] if transcription else ''
            t_words = extract_content_words(clean_stimulus(str(stimulus)) if stimulus else '')
            r_words = extract_content_words(clean_transcription(str(transcription)) if transcription else '')
            _, _, ratio = fuzzy_word_overlap(t_words, r_words) if t_words else (0,0,0)

            print(f"{sentence_no:<5} {score:<7} {ratio*100:<8.1f}% {clean_s:<35} {clean_t}")

        total = sum(scores)
        avg = total / len(scores) if scores else 0
        results_summary[sheet_name] = {
            'scores': scores,
            'total': total,
            'average': avg,
            'distribution': {i: scores.count(i) for i in range(5)}
        }
        print(f"\n  Total Score: {total}/120  |  Average: {avg:.2f}/4  |  Distribution: {results_summary[sheet_name]['distribution']}")

    wb.save(output_path)
    print(f"\n\n✅ Scored workbook saved to: {output_path}")
    return results_summary


# ─────────────────────────────────────────────
# SECTION 6: MAIN
# ─────────────────────────────────────────────

if __name__ == "__main__":
    INPUT_FILE  = '/mnt/user-data/uploads/AutoEIT_Sample_Transcriptions_for_Scoring.xlsx'
    OUTPUT_FILE = '/home/claude/AutoEIT_Scored_Output.xlsx'

    print("AutoEIT - Automated Scoring System")
    print("Anand S | GSoC 2026 | HumanAI Foundation")
    print("=" * 55)

    summary = process_workbook(INPUT_FILE, OUTPUT_FILE)

    print("\n\n📊 SUMMARY ACROSS ALL PARTICIPANTS")
    print("=" * 55)
    for participant, data in summary.items():
        print(f"\n{participant}:")
        print(f"  Total   : {data['total']}/120")
        print(f"  Average : {data['average']:.2f} / 4.00")
        print(f"  Score Distribution (0→4): {data['distribution']}")

    # Show overall corpus stats
    all_scores = [s for d in summary.values() for s in d['scores']]
    print(f"\n📈 Corpus-wide average score: {sum(all_scores)/len(all_scores):.2f} / 4.00")
    print(f"   Total sentences scored : {len(all_scores)}")
